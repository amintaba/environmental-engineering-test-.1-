
import openpyxl

Tabel_clim = openpyxl.load_workbook("Table_S1.xlsx")

Gauge_list = Tabel_clim["Sheet1"]

Gauge_per_NS = {}

print(Gauge_list.max_row)

for Gauge_row  in range(4 ,Gauge_list.max_row +1):
    NS_name = Gauge_list.cell(Gauge_row , 7).value
    R =  Gauge_list.cell(Gauge_row , 6).value
    fee = Gauge_list.cell(Gauge_row, 8).value

    #calculation number of Gauge per NS & R & FEE
    if NS_name in Gauge_per_NS:
        NS_num = Gauge_per_NS.get(NS_name , R , fee)
        Gauge_per_NS[NS_name , R , fee] = NS_num +1

    else:
        print("adding a new NS")
        Gauge_per_NS[NS_name , R , fee] =  Gauge_list.cell(Gauge_row ,1).value



print(Gauge_per_NS)
